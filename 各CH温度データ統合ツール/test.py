import sys
import pandas as pd
import os
import re
from PyQt6.QtWidgets import QApplication, QFileDialog
from openpyxl import load_workbook

# ファイル選択ダイアログ
app = QApplication(sys.argv)

def select_files(title="ファイルを選択", multi=True):
    options = QFileDialog.Option.ReadOnly
    if multi:
        files, _ = QFileDialog.getOpenFileNames(None, title, "", "CSV or Excel Files (*.csv *.xlsx *.xls)")
    else:
        files, _ = QFileDialog.getOpenFileName(None, title, "", "CSV or Excel Files (*.csv *.xlsx *.xls)")
    return files if multi else [files]

# 元ファイル選択
source_files = select_files("統合する元ファイルを選択 (複数選択可)", multi=True)
if not source_files:
    sys.exit("元ファイルが選択されませんでした")

# 統合先ファイル選択
target_file = select_files("統合先のExcelファイルを選択", multi=False)[0]
if not target_file:
    sys.exit("統合先ファイルが選択されませんでした")

target_wb = load_workbook(target_file)

# CH番号を抽出する関数
def extract_ch_number(filename):
    match = re.search(r"CH(\d+)", filename)  # CHに続く数字を検索
    return match.group(1) if match else None

# 温度順に並び替え
def get_temperature_order(filename):
    if "10度" in filename:
        return 1
    elif "25度" in filename:
        return 2
    elif "40度" in filename:
        return 3
    return 99  # 不明な場合は最後

source_files.sort(key=get_temperature_order)

# データ統合
for source_file in source_files:
    filename = os.path.basename(source_file)
    ch_number = extract_ch_number(filename)
    
    if ch_number is None:
        print(f"スキップ: {filename}（CH番号が見つかりませんでした）")
        continue
    
    sheet_name = f"CH{ch_number}"
    
    if sheet_name not in target_wb.sheetnames:
        print(f"スキップ: {filename}（{sheet_name} シートがありません）")
        continue

    # 温度順を取得
    temp_order = get_temperature_order(filename)
    if temp_order == 99:
        print(f"スキップ: {filename}（温度情報が見つかりませんでした）")
        continue

    # 書き込み開始列を決定
    col_mapping = {1: 1, 2: 7, 3: 13}  # 10度=1列目, 25度=7列目, 40度=13列目（1列追加）
    vrd_col_mapping = {1: 20, 2: 26, 3: 32}  # 10度=T列, 25度=Z列, 40度=AF列
    start_col = col_mapping[temp_order]  # 各温度データを横方向に配置
    vrd_start_col = vrd_col_mapping[temp_order]  # Vrd=2のデータの開始列
    
    # データ読み込み
    if filename.endswith(".csv"):
        df = pd.read_csv(source_file, sep=None, engine="python", skiprows=9)  # 先頭9行をスキップ
    else:
        df = pd.read_excel(source_file)
    
    # Id(mA) 列を追加
    if "Id (A)" in df.columns:
        df.insert(df.columns.get_loc("Id (A)") + 1, "Id (mA)", df["Id (A)"] * 1000)
    
    # 統合先シート取得
    sheet = target_wb[sheet_name]
    
    # 1行目に温度情報のみを追加
    temp_labels = {1: "10度", 2: "25度", 3: "40度"}
    sheet.cell(row=1, column=start_col, value=temp_labels[temp_order])
    
    # 2行目にヘッダーを追加
    headers = list(df.columns)
    for c_idx, header in enumerate(headers):
        sheet.cell(row=2, column=start_col + c_idx, value=header)
    
    # データの貼り付け（各温度データを横方向に配置、2行目以降にデータを記入）
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=0):  # 0列目からすべてコピー
            sheet.cell(row=r_idx, column=start_col + c_idx, value=value)
    
    # Vrd(V)が2Vになっている範囲(A列～R列)を取得し、各温度ごとに異なる列にペースト
    if "Vrd (V)" in df.columns:
        vrd_2_rows = df[df["Vrd (V)"] == 2].index
        if not vrd_2_rows.empty:
            start_idx = vrd_2_rows[0]
            end_idx = start_idx
            while end_idx + 1 in vrd_2_rows:
                end_idx += 1
            selected_data = df.iloc[start_idx:end_idx + 1, :18]  # A列～R列
            for r_idx, row in enumerate(selected_data.itertuples(index=False), start=3):
                for c_idx, value in enumerate(row, start=0):
                    sheet.cell(row=r_idx, column=vrd_start_col + c_idx, value=value)  # T, Z, AF列にデータ配置
    
    print(f"統合完了: {filename} → {sheet_name}（開始セル: {start_col}1）")

target_wb.save(target_file)
target_wb.close()
print("統合処理が完了しました！")
